"""Export classifica to .xlsx.

Fogli prodotti (in ordine):
  1. Assoluta
  2. Uomini (classifica di sesso M)
  3. Donne  (classifica di sesso F)
  4. Una sheet per ogni categoria (es. M18-24, M25-29 … F18-24 …)
"""

import os
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from logic.classifica import RigaClassifica, calcola_classifica
from utils.paths import get_export_dir

_STATO_LABELS = {"ok": "OK", "dsq": "DSQ", "dnf": "DNF", "dns": "DNS", "": "—"}

# Header style
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_HDR_FILL = PatternFill("solid", fgColor="2563EB")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center")

# Even row fill
_EVEN_FILL = PatternFill("solid", fgColor="EFF6FF")

# Grayed text for non-ok rows
_GRAY_FONT = Font(color="94A3B8")

_THIN = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_row(ws, headers: list[str], col_widths: list[int]) -> None:
    ws.append(headers)
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20


def _data_row(ws, row_idx: int, values: list, grayed: bool = False) -> None:
    fill = _EVEN_FILL if row_idx % 2 == 0 else None
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(
            horizontal="center" if col_idx != 3 else "left",
            vertical="center",
        )
        cell.border = _BORDER
        if grayed:
            cell.font = _GRAY_FONT
        if fill:
            cell.fill = fill
    ws.row_dimensions[row_idx].height = 16


def _sheet_assoluta(wb: openpyxl.Workbook, righe: list[RigaClassifica]) -> None:
    ws = wb.active
    ws.title = "Assoluta"
    ws.freeze_panes = "A2"

    headers = ["Pos.", "Pett.", "Atleta", "Cat.", "Sesso", "Tempo", "Stato"]
    widths  = [6,       8,       28,       12,     8,       12,      8]
    _header_row(ws, headers, widths)

    sorted_rows = sorted(
        righe,
        key=lambda r: (0 if r.pos_assoluta else 1, r.pos_assoluta or 9999),
    )
    for i, r in enumerate(sorted_rows, start=2):
        grayed = r.stato in ("dsq", "dnf", "dns")
        _data_row(ws, i, [
            r.pos_assoluta if r.pos_assoluta else "—",
            r.pettorale,
            r.nome_atleta,
            r.categoria,
            r.sesso,
            r.tempo_display,
            _STATO_LABELS.get(r.stato, r.stato or "—"),
        ], grayed=grayed)


def _sheet_categoria(wb: openpyxl.Workbook, nome_cat: str, righe_cat: list[RigaClassifica]) -> None:
    """Crea un foglio dedicato a una singola categoria."""
    # I nomi dei fogli Excel non possono superare 31 caratteri
    ws = wb.create_sheet(nome_cat[:31])
    ws.freeze_panes = "A2"

    headers = ["Pos.", "Pett.", "Atleta", "Tempo", "Stato"]
    widths  = [6,       8,       32,       12,      8]
    _header_row(ws, headers, widths)

    sorted_rows = sorted(
        righe_cat,
        key=lambda r: (0 if r.pos_categoria else 1, r.pos_categoria or 9999),
    )
    for i, r in enumerate(sorted_rows, start=2):
        grayed = r.stato in ("dsq", "dnf", "dns")
        _data_row(ws, i, [
            r.pos_categoria if r.pos_categoria else "—",
            r.pettorale,
            r.nome_atleta,
            r.tempo_display,
            _STATO_LABELS.get(r.stato, r.stato or "—"),
        ], grayed=grayed)


def _sheet_per_sesso(wb: openpyxl.Workbook, righe: list[RigaClassifica], sesso: str, title: str) -> None:
    ws = wb.create_sheet(title)
    ws.freeze_panes = "A2"

    headers = ["Pos.", "Pett.", "Atleta", "Cat.", "Tempo", "Stato"]
    widths  = [6,       8,       28,       12,     12,      8]
    _header_row(ws, headers, widths)

    filtered = sorted(
        [r for r in righe if r.sesso.upper() in (sesso.upper(), )],
        key=lambda r: (0 if r.pos_sesso else 1, r.pos_sesso or 9999),
    )
    for i, r in enumerate(filtered, start=2):
        grayed = r.stato in ("dsq", "dnf", "dns")
        _data_row(ws, i, [
            r.pos_sesso if r.pos_sesso else "—",
            r.pettorale,
            r.nome_atleta,
            r.categoria,
            r.tempo_display,
            _STATO_LABELS.get(r.stato, r.stato or "—"),
        ], grayed=grayed)


def esporta_xlsx(
    righe: list[RigaClassifica],
    nome_gara: str,
    nome_evento: str = "",
    dest_path: Optional[str] = None,
) -> str:
    """Build workbook and write to dest_path (or auto-named in export dir). Returns path."""
    if not dest_path:
        safe = "".join(c if c.isalnum() or c in "-_ " else "_" for c in nome_gara)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest_path = os.path.join(get_export_dir(), f"classifica_{safe}_{ts}.xlsx")

    wb = openpyxl.Workbook()

    _sheet_assoluta(wb, righe)
    _sheet_per_sesso(wb, righe, "M", "Uomini")
    _sheet_per_sesso(wb, righe, "F", "Donne")

    # Un foglio per ogni categoria, nello stesso ordine dell'app
    categorie = sorted({r.categoria for r in righe if r.categoria})
    for nome_cat in categorie:
        righe_cat = [r for r in righe if r.categoria == nome_cat]
        _sheet_categoria(wb, nome_cat, righe_cat)

    # Metadata
    wb.properties.title = f"Classifica — {nome_gara}"
    if nome_evento:
        wb.properties.subject = nome_evento
    wb.properties.creator = "UfensCrono"

    wb.save(dest_path)
    return dest_path
