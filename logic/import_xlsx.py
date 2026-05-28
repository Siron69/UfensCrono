import re
import sqlite3
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Optional

import openpyxl

from models.atleta import Atleta
from models.gara import Categoria
from db.queries import atleti as qa
from db.queries import gare as qg
from db.queries.eventi import get_by_id as get_evento
from logic.categorie import calcola_categoria


@dataclass
class ImportResult:
    n_inseriti: int = 0
    n_aggiornati: int = 0
    n_iscrizioni: int = 0
    n_saltati: int = 0
    errori: list[tuple[int, str]] = field(default_factory=list)

    def __str__(self) -> str:
        lines = [
            f"Atleti inseriti:  {self.n_inseriti}",
            f"Atleti aggiornati: {self.n_aggiornati}",
            f"Iscrizioni create: {self.n_iscrizioni}",
            f"Saltati:           {self.n_saltati}",
        ]
        if self.errori:
            lines.append(f"Errori:            {len(self.errori)}")
        return "\n".join(lines)


def _normalize_col(header) -> str:
    """Normalizza un'intestazione di colonna per il confronto robusto.

    Gestisce: BOM, spazi non-breaking, accenti in forma NFC/NFD,
    maiuscole/minuscole, spazi ridondanti.
    """
    if header is None:
        return ''
    s = str(header)
    s = s.replace('﻿', '')       # BOM
    s = s.replace(' ', ' ')      # spazio non-breaking → spazio normale
    s = unicodedata.normalize('NFC', s)  # à = a+combining → à (precomposta)
    s = ' '.join(s.split())              # collassa spazi multipli/tab
    return s.lower()


# Mapping nome colonna (lowercase, normalizzato) → (campo_db, tabella)
_COL_MAP: dict[str, tuple[str, str]] = {
    'id':                   ('source_id',        'atleti'),
    'order id':             ('source_order_id',   'atleti'),
    'nome':                 ('nome',              'atleti'),
    'cognome':              ('cognome',           'atleti'),
    'sesso':                ('sesso',             'atleti'),
    'data nascita':         ('data_nascita',      'atleti'),
    'luogo di nascita':     ('luogo_nascita',     'atleti'),
    'nazionalità':     ('nazionalita',       'atleti'),   # à precomposta
    'nazionalita':          ('nazionalita',       'atleti'),
    'codice fiscale':       ('codice_fiscale',    'atleti'),
    'tessera':              ('tessera',           'atleti'),
    'tessera2':             ('tessera2',          'atleti'),
    'ente':                 ('ente',              'atleti'),
    'codice società':  ('codice_societa',    'atleti'),   # à precomposta
    'codice societa':       ('codice_societa',    'atleti'),
    'società':         ('societa',           'atleti'),   # à precomposta
    'societa':              ('societa',           'atleti'),
    'categoria':            ('categoria',         'atleti'),
    'scadenza certificato': ('scad_certificato',  'atleti'),
    'stato certificato':    ('stato_cert',        'atleti'),
    'telefono':             ('telefono',          'atleti'),
    'cellulare':            ('cellulare',         'atleti'),
    'e-mail':               ('email',             'atleti'),
    'email':                ('email',             'atleti'),
    'pettorale':            ('pettorale',         'iscrizioni'),
    'pettorale circuito':   ('pettorale_circ',    'iscrizioni'),
    'codice chip':          ('codice_chip',       'iscrizioni'),
    'quota':                ('quota',             'iscrizioni'),
    'stato lw':             ('stato_lw',          'iscrizioni'),
    # ── Alias comuni (varianti di nome colonna) ──────────────────────────
    'first name':           ('nome',              'atleti'),
    'last name':            ('cognome',           'atleti'),
    'surname':              ('cognome',           'atleti'),
    'firstname':            ('nome',              'atleti'),
    'lastname':             ('cognome',           'atleti'),
    'data di nascita':      ('data_nascita',      'atleti'),
    'date of birth':        ('data_nascita',      'atleti'),
    'dob':                  ('data_nascita',      'atleti'),
    'birth date':           ('data_nascita',      'atleti'),
    'soc':                  ('societa',           'atleti'),
    'società sportiva':     ('societa',           'atleti'),
    'associazione':         ('societa',           'atleti'),
    'club':                 ('societa',           'atleti'),
    'gender':               ('sesso',             'atleti'),
    'sex':                  ('sesso',             'atleti'),
    'bib':                  ('pettorale',         'iscrizioni'),
    'bib number':           ('pettorale',         'iscrizioni'),
    'numero pettorale':     ('pettorale',         'iscrizioni'),
    'chip':                 ('codice_chip',       'iscrizioni'),
    'cf':                   ('codice_fiscale',    'atleti'),
    'fiscal code':          ('codice_fiscale',    'atleti'),
}


# Pre-normalizza le chiavi del dizionario una volta sola
_COL_MAP = {_normalize_col(k): v for k, v in _COL_MAP.items()}


def _str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def normalizza_sesso(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    v = val.strip().upper()
    if v in ('M', 'MASCHIO', 'MALE', 'H', 'HOMME', 'UOMO'):
        return 'M'
    if v in ('F', 'FEMMINA', 'FEMALE', 'DONNA', 'W', 'WOMAN'):
        return 'F'
    return None


def normalizza_data(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    # DD/MM/YYYY
    m = re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02}-{int(d):02}"
    # DD-MM-YYYY
    m = re.fullmatch(r'(\d{1,2})-(\d{1,2})-(\d{4})', s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02}-{int(d):02}"
    # YYYY-MM-DD
    if re.fullmatch(r'\d{4}-\d{2}-\d{2}', s):
        return s
    return None


def riconosci_colonne(headers: list[str]) -> list[bool]:
    """Ritorna una lista di bool: True se la colonna è mappata, False altrimenti."""
    return [_normalize_col(h) in _COL_MAP for h in headers]


def leggi_xlsx_preview(path: str) -> tuple[list[str], list[list[str]]]:
    """Restituisce (intestazioni, prime 5 righe dati) come stringhe."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True, max_row=6))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append([str(v).strip() if v is not None else '' for v in row])
    return headers, data_rows


def importa_xlsx(
    path: str,
    gara_id: int,
    conn: sqlite3.Connection,
) -> ImportResult:
    result = ImportResult()

    # Carica workbook
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return result

    # Mappa indici colonne
    raw_headers = all_rows[0]
    col_index: dict[str, tuple[int, str]] = {}  # field_name → (col_i, table)
    for i, h in enumerate(raw_headers):
        if h is None:
            continue
        key = _normalize_col(h)
        if key in _COL_MAP:
            field_name, table = _COL_MAP[key]
            col_index[field_name] = (i, table)

    # Info gara per categoria e validazione cross-gara
    gara = qg.get_by_id(conn, gara_id)
    evento_id: Optional[int] = gara.evento_id if gara else None
    categorie: list[Categoria] = qg.get_categorie(conn, gara_id)
    anno_gara = 2025
    if gara:
        ev = get_evento(conn, gara.evento_id)
        if ev and ev.data:
            try:
                anno_gara = int(ev.data.split('-')[0])
            except Exception:
                pass

    already_registered = qg.atleti_iscritti_ids(conn, gara_id)

    def _get(row, field_name):
        info = col_index.get(field_name)
        if info is None:
            return None
        i, _ = info
        return row[i] if i < len(row) else None

    for row_num, row in enumerate(all_rows[1:], start=2):
        try:
            nome = _str(_get(row, 'nome'))
            cognome = _str(_get(row, 'cognome'))
            if not nome or not cognome:
                result.errori.append((row_num, "Nome o cognome mancante"))
                continue

            sesso = normalizza_sesso(_str(_get(row, 'sesso')))
            if not sesso:
                result.errori.append((row_num, f"Sesso non valido: {_get(row, 'sesso')!r}"))
                continue

            dob = normalizza_data(_get(row, 'data_nascita'))
            if not dob:
                result.errori.append((row_num, f"Data nascita non valida: {_get(row, 'data_nascita')!r}"))
                continue

            cf_raw = _str(_get(row, 'codice_fiscale'))
            cf = cf_raw.upper() if cf_raw else None

            atleta = Atleta(
                nome=nome,
                cognome=cognome,
                sesso=sesso,
                data_nascita=dob,
                luogo_nascita=_str(_get(row, 'luogo_nascita')),
                nazionalita=_str(_get(row, 'nazionalita')) or 'ITA',
                codice_fiscale=cf,
                societa=_str(_get(row, 'societa')),
                codice_societa=_str(_get(row, 'codice_societa')),
                tessera=_str(_get(row, 'tessera')),
                tessera2=_str(_get(row, 'tessera2')),
                ente=_str(_get(row, 'ente')),
                categoria=_str(_get(row, 'categoria')),
                scad_certificato=normalizza_data(_get(row, 'scad_certificato')),
                stato_cert=_str(_get(row, 'stato_cert')),
                telefono=_str(_get(row, 'telefono')),
                cellulare=_str(_get(row, 'cellulare')),
                email=_str(_get(row, 'email')),
                source_id=_str(_get(row, 'source_id')),
                source_order_id=_str(_get(row, 'source_order_id')),
            )

            # Deduplicazione
            existing = None
            if cf:
                existing = qa.find_by_cf(conn, cf)
            if existing is None:
                existing = qa.find_by_nome_cognome_dob(conn, nome, cognome, dob)

            if existing:
                atleta.id = existing.id
                qa.update_from_import(conn, atleta)
                result.n_aggiornati += 1
                atleta_id = existing.id
            else:
                atleta_id = qa.insert(conn, atleta)
                result.n_inseriti += 1

            # Iscrizione
            if atleta_id in already_registered:
                result.n_saltati += 1
                continue

            pettorale = _str(_get(row, 'pettorale'))
            if not pettorale:
                result.n_saltati += 1
                continue

            # Validazione unicità pettorale cross-gara
            if evento_id:
                gara_conflitto = qg.get_pettorale_conflitto(conn, evento_id, pettorale, exclude_gara_id=gara_id)
                if gara_conflitto:
                    result.errori.append((
                        row_num,
                        f"Pettorale {pettorale!r} già usato nella gara «{gara_conflitto}» dello stesso evento",
                    ))
                    continue

            cat_calc = calcola_categoria(categorie, dob, sesso, anno_gara)

            try:
                qg.add_iscrizione(
                    conn, gara_id, atleta_id, pettorale,
                    categoria_calc=cat_calc,
                )
                already_registered.add(atleta_id)
                result.n_iscrizioni += 1
            except Exception as e:
                result.errori.append((row_num, f"Iscrizione fallita: {e}"))

        except Exception as e:
            result.errori.append((row_num, str(e)))

    return result
