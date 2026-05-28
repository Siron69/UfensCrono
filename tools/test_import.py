"""
Strumento diagnostico CLI per il debug dell'import XLSX.

Uso:
    cd <root progetto>
    .venv\Scripts\python.exe tools\test_import.py <percorso_file.xlsx> [numero_foglio]

Mostra:
  - Fogli disponibili nel workbook
  - Intestazioni raw + repr() + forma normalizzata
  - Riconoscimento colonne (match con _COL_MAP)
  - col_index che verrebbe costruito durante l'import
  - Prime 3 righe di dati
"""

import sys
import os
import io

# Forza stdout UTF-8 su Windows per evitare errori cp1252 con caratteri Unicode
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Assicura che il root del progetto sia nel path
ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), '..'))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import unicodedata
import openpyxl

from logic.import_xlsx import _normalize_col, _COL_MAP


def _hr(char: str = '─', width: int = 70) -> None:
    print(char * width)


def analizza_xlsx(path: str, sheet_index: int = 0) -> None:
    # ── Carica workbook ───────────────────────────────────────────────────
    print(f"\nFile: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Fogli disponibili ({len(sheet_names)}): {sheet_names}")

    if sheet_index >= len(sheet_names):
        print(f"Indice foglio {sheet_index} fuori range. Uso foglio 0.")
        sheet_index = 0

    ws = wb.worksheets[sheet_index]
    print(f"Foglio selezionato: [{sheet_index}] '{sheet_names[sheet_index]}'")
    _hr()

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        print("ERRORE: il foglio è vuoto!")
        return

    raw_headers = all_rows[0]
    n_cols = len(raw_headers)
    print(f"Numero di colonne: {n_cols}")
    _hr()

    # ── Analisi intestazioni ──────────────────────────────────────────────
    print(f"{'Col':>4}  {'Raw header':<30}  {'repr()':<45}  {'Normalizzato':<30}  {'Mappato a'}")
    _hr()

    col_index: dict = {}
    n_riconosciute = 0

    for i, h in enumerate(raw_headers):
        raw = str(h) if h is not None else '(None)'
        norm = _normalize_col(h) if h is not None else ''
        match = _COL_MAP.get(norm)
        if match:
            field_name, table = match
            col_index[field_name] = (i, table)
            mappato = f"→ {table}.{field_name}"
            n_riconosciute += 1
        else:
            mappato = '— non riconosciuto'

        # Mostra eventuali caratteri speciali nel repr
        r = repr(h)
        print(f"{i:>4}  {raw:<30}  {r:<45}  {norm:<30}  {mappato}")

    _hr()
    print(f"Colonne riconosciute: {n_riconosciute}/{n_cols}")
    _hr()

    # ── col_index risultante ──────────────────────────────────────────────
    print("\ncol_index costruito (campo → (indice_colonna, tabella)):")
    if col_index:
        for field, (idx, tbl) in col_index.items():
            raw_val = str(raw_headers[idx]) if raw_headers[idx] is not None else '(None)'
            print(f"  {field:<25} col {idx:>3}  [{raw_val}]  ({tbl})")
    else:
        print("  (vuoto — nessuna colonna riconosciuta!)")

    # ── Campi obbligatori ────────────────────────────────────────────────
    _hr()
    obbligatori = ['nome', 'cognome', 'sesso', 'data_nascita']
    print("Campi obbligatori:")
    for f in obbligatori:
        stato = "✓ presente" if f in col_index else "✗ MANCANTE"
        print(f"  {f:<20} {stato}")

    # ── Prime 3 righe dati ───────────────────────────────────────────────
    _hr()
    print("\nPrime 3 righe dati:")
    for row_num, row in enumerate(all_rows[1:4], start=2):
        print(f"\n  Riga {row_num}:")
        for field_name, (col_i, _) in col_index.items():
            val = row[col_i] if col_i < len(row) else '(fuori range)'
            print(f"    {field_name:<25} = {val!r}")

    print()


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    path = sys.argv[1]
    sheet_index = int(sys.argv[2]) if len(sys.argv) > 2 else 0

    if not os.path.isfile(path):
        print(f"Errore: file non trovato: {path}")
        sys.exit(1)

    analizza_xlsx(path, sheet_index)


if __name__ == '__main__':
    main()
