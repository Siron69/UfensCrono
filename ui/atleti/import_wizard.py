from typing import Optional

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QStackedWidget,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QComboBox, QTextEdit, QFileDialog, QMessageBox, QHeaderView,
    QProgressBar, QWidget,
)
from PyQt6.QtCore import Qt

from db.connection import get_connection
from db.queries import eventi as qev
from db.queries import gare as qg
from logic.import_xlsx import leggi_xlsx_preview, importa_xlsx


class ImportWizard(QDialog):
    def __init__(self, parent=None, gara_id: Optional[int] = None):
        super().__init__(parent)
        self._gara_id_preset = gara_id   # pre-selezionato se aperto da iscrizioni
        self._file_path: Optional[str] = None
        self._selected_gara_id: Optional[int] = None
        self.setWindowTitle("Importa atleti da XLSX")
        self.setMinimumSize(820, 560)
        self._build_ui()
        if gara_id:
            self._preselect_gara(gara_id)

    # ── Costruzione UI ────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)

        # Indicatore step
        self.lbl_step = QLabel()
        self.lbl_step.setStyleSheet("color: #64748b; font-size: 12px;")
        layout.addWidget(self.lbl_step)

        self.stack = QStackedWidget()
        self.stack.addWidget(self._build_step1())
        self.stack.addWidget(self._build_step2())
        self.stack.addWidget(self._build_step3())
        layout.addWidget(self.stack, stretch=1)

        # Navigazione
        nav = QHBoxLayout()
        self.btn_back = QPushButton("← Indietro")
        self.btn_back.clicked.connect(self._on_back)
        nav.addWidget(self.btn_back)
        nav.addStretch()
        self.btn_cancel = QPushButton("Annulla")
        self.btn_cancel.clicked.connect(self.reject)
        nav.addWidget(self.btn_cancel)
        self.btn_next = QPushButton("Avanti →")
        self.btn_next.clicked.connect(self._on_next)
        nav.addWidget(self.btn_next)
        layout.addLayout(nav)

        self._goto_step(0)

    def _build_step1(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)

        lay.addWidget(QLabel("<b>Passo 1 — Seleziona file XLSX</b>"))

        file_row = QHBoxLayout()
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setReadOnly(True)
        self.file_path_edit.setPlaceholderText("Nessun file selezionato…")
        file_row.addWidget(self.file_path_edit, stretch=1)
        btn_browse = QPushButton("Sfoglia…")
        btn_browse.clicked.connect(self._on_browse)
        file_row.addWidget(btn_browse)
        lay.addLayout(file_row)

        lay.addWidget(QLabel("Anteprima (prime 5 righe):"))
        self.preview_table = QTableWidget(0, 0)
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.verticalHeader().setVisible(False)
        self.preview_table.setAlternatingRowColors(True)
        lay.addWidget(self.preview_table, stretch=1)

        self.lbl_n_righe = QLabel("")
        self.lbl_n_righe.setStyleSheet("color: #64748b;")
        lay.addWidget(self.lbl_n_righe)

        return w

    def _build_step2(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.addWidget(QLabel("<b>Passo 2 — Seleziona gara di destinazione</b>"))
        lay.addSpacing(12)

        from PyQt6.QtWidgets import QFormLayout
        form = QFormLayout()

        self.combo_evento = QComboBox()
        self.combo_evento.currentIndexChanged.connect(self._on_evento_changed)
        form.addRow("Evento:", self.combo_evento)

        self.combo_gara = QComboBox()
        self.combo_gara.currentIndexChanged.connect(self._update_nav)
        form.addRow("Gara:", self.combo_gara)

        lay.addLayout(form)
        lay.addStretch()

        self.lbl_gara_info = QLabel("")
        self.lbl_gara_info.setStyleSheet("color: #64748b;")
        lay.addWidget(self.lbl_gara_info)

        return w

    def _build_step3(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.addWidget(QLabel("<b>Passo 3 — Conferma e importa</b>"))

        self.lbl_riepilogo = QLabel("")
        self.lbl_riepilogo.setWordWrap(True)
        lay.addWidget(self.lbl_riepilogo)

        self.btn_importa = QPushButton("Importa")
        self.btn_importa.setStyleSheet(
            "background-color: #2563eb; color: white; padding: 8px 24px;"
            "font-size: 14px; border-radius: 4px;"
        )
        self.btn_importa.clicked.connect(self._on_importa)
        lay.addWidget(self.btn_importa, alignment=Qt.AlignmentFlag.AlignLeft)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        lay.addWidget(self.progress)

        lay.addWidget(QLabel("Risultato:"))
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setMaximumHeight(200)
        lay.addWidget(self.result_text)

        return w

    # ── Navigazione step ──────────────────────────────────────────────────

    def _goto_step(self, step: int) -> None:
        self._step = step
        self.stack.setCurrentIndex(step)
        labels = [
            "Passo 1 di 3: Selezione file",
            "Passo 2 di 3: Selezione gara",
            "Passo 3 di 3: Conferma",
        ]
        self.lbl_step.setText(labels[step])
        self.btn_back.setEnabled(step > 0)
        self._update_nav()

        if step == 1:
            self._populate_eventi()
        if step == 2:
            gid = self._selected_gara_id
            if gid:
                gara = qg.get_by_id(get_connection(), gid)
                n = self._count_file_rows()
                self.lbl_riepilogo.setText(
                    f"File: <b>{self._file_path}</b><br>"
                    f"Gara: <b>{gara.nome if gara else ''}</b><br>"
                    f"Righe nel file: <b>{n}</b><br><br>"
                    "Premi «Importa» per avviare l'importazione."
                )

    def _on_next(self) -> None:
        if self._step == 0:
            if not self._file_path:
                QMessageBox.warning(self, "File mancante", "Seleziona prima un file XLSX.")
                return
            # Se gara preset salta step 2
            if self._gara_id_preset:
                self._selected_gara_id = self._gara_id_preset
                self._goto_step(2)
            else:
                self._goto_step(1)
        elif self._step == 1:
            gid = self.combo_gara.currentData()
            if not gid:
                QMessageBox.warning(self, "Gara mancante", "Seleziona una gara di destinazione.")
                return
            self._selected_gara_id = gid
            self._goto_step(2)

    def _on_back(self) -> None:
        if self._step == 2 and self._gara_id_preset:
            self._goto_step(0)
        elif self._step > 0:
            self._goto_step(self._step - 1)

    def _update_nav(self) -> None:
        step = getattr(self, '_step', 0)
        if step == 0:
            self.btn_next.setText("Avanti →")
            self.btn_next.setEnabled(bool(self._file_path))
        elif step == 1:
            self.btn_next.setText("Avanti →")
            self.btn_next.setEnabled(bool(self.combo_gara.currentData()))
        else:
            self.btn_next.setText("Chiudi")
            self.btn_next.clicked.disconnect()
            self.btn_next.clicked.connect(self.accept)

    # ── Step 1: selezione file ────────────────────────────────────────────

    def _on_browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Seleziona file XLSX", "", "File Excel (*.xlsx *.xls)"
        )
        if not path:
            return
        try:
            headers, rows = leggi_xlsx_preview(path)
        except Exception as e:
            QMessageBox.critical(self, "Errore lettura file", str(e))
            return

        self._file_path = path
        self.file_path_edit.setText(path)

        # Conta righe reali
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        total = wb.active.max_row - 1  # escludi header
        wb.close()
        self.lbl_n_righe.setText(f"Righe totali nel file: {total}")

        # Preview
        n_cols = len(headers)
        self.preview_table.setColumnCount(n_cols)
        self.preview_table.setHorizontalHeaderLabels(headers)
        self.preview_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                self.preview_table.setItem(r, c, QTableWidgetItem(val))
        self.preview_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        self._update_nav()

    def _count_file_rows(self) -> int:
        if not self._file_path:
            return 0
        import openpyxl
        wb = openpyxl.load_workbook(self._file_path, read_only=True)
        n = (wb.active.max_row or 1) - 1
        wb.close()
        return n

    # ── Step 2: selezione gara ────────────────────────────────────────────

    def _populate_eventi(self) -> None:
        conn = get_connection()
        eventi = qev.get_all(conn)
        self.combo_evento.blockSignals(True)
        self.combo_evento.clear()
        self.combo_evento.addItem("— Seleziona evento —", None)
        for ev in eventi:
            self.combo_evento.addItem(f"{ev.nome} ({ev.data})", ev.id)
        self.combo_evento.blockSignals(False)
        self._on_evento_changed()

    def _on_evento_changed(self) -> None:
        eid = self.combo_evento.currentData()
        self.combo_gara.clear()
        self.combo_gara.addItem("— Seleziona gara —", None)
        if eid:
            gare = qg.get_by_evento(get_connection(), eid)
            for g in gare:
                if g.stato == 'bozza':
                    self.combo_gara.addItem(g.nome, g.id)
        self._update_nav()

    def _preselect_gara(self, gara_id: int) -> None:
        gara = qg.get_by_id(get_connection(), gara_id)
        if gara:
            self._selected_gara_id = gara_id
            self.lbl_gara_info.setText(
                f"Gara preselezionata: <b>{gara.nome}</b>"
            )

    # ── Step 3: importa ───────────────────────────────────────────────────

    def _on_importa(self) -> None:
        if not self._file_path or not self._selected_gara_id:
            return

        self.btn_importa.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setRange(0, 0)  # indeterminato

        try:
            result = importa_xlsx(
                self._file_path,
                self._selected_gara_id,
                get_connection(),
            )
        except Exception as e:
            self.result_text.setPlainText(f"Errore durante l'import:\n{e}")
            self.progress.setVisible(False)
            self.btn_importa.setEnabled(True)
            return

        self.progress.setVisible(False)
        lines = [str(result)]
        if result.errori:
            lines.append("\nDettaglio errori:")
            for row_num, msg in result.errori[:20]:
                lines.append(f"  Riga {row_num}: {msg}")
            if len(result.errori) > 20:
                lines.append(f"  … e altri {len(result.errori) - 20} errori")
        self.result_text.setPlainText("\n".join(lines))
        self.btn_importa.setEnabled(False)
        self._update_nav()
